import sys
import os
from config import *
from utility import *
from scapy.all import *
import openpyxl
import scipy.io as scio
from scipy.special import gamma, hyp2f1, betainc
import random
import time
import numpy as np
from ast import literal_eval
import logging
import logging.handlers

my_logger = logging.getLogger('MyLogger')
my_logger.setLevel(logging.DEBUG)

handler = logging.handlers.SysLogHandler(address = '/dev/log')

my_logger.addHandler(handler)


class Ground_station:
    def __init__(self, id, name, lat, long, alt):
        self.id = id
        self.name = name
        self.position = (lat, long, alt)
        self.candidates = None
        self.current_satellite = None

    def get_position(self):
        return self.position
    
    def print_info(self):
        print(f"Ground station ({self.id}) {self.name} is located on {self.position}.")

    def set_candidates(self, sat_list):
        self.candidates = sat_list

    def set_current_satellite(self, sat):
        # if self.current_satellite is None:
        #     print(f"GS {self.id} is being connected to satellite {sat.id}...")
        # else:
        #     print(f"GS {self.id} is being connected to satellite {sat.id} from satellite {self.current_satellite.id}")
        self.current_satellite = sat

    def get_current_satellite(self):
        return self.current_satellite
        
class Satellite:
    def __init__(self, id, name, positions):
        self.id = id
        self.name = name
        self.positions = positions
        self.current_gs = list()
    
    def print_info(self, t):
        print(f"Satellite ({self.id}) {self.name} is on {self.get_position(t)} at time {t}.")

    def get_position(self, curr_time):
        return self.positions[curr_time]
    
    def set_current_gs(self, gs):
        self.current_gs.append(gs)
    

# find a satellite with minimum link loss value among candidate satellites
def select_next_satellite(candidates, observer_pos, curr_time):  
    min_satellite = None
    min_loss = float('inf')

    for sat in candidates:
        link_loss = calculate_link_loss(sat.get_position(curr_time), observer_pos)
        if link_loss < min_loss:
            min_loss = link_loss
            min_satellite = sat

    return min_satellite, min_loss

def calculate_freeze_duration():
    ho_trigger_prob = calc_trigger_prob(ho_threshold, m)
    ho_failure_prob = calc_failure_prob(hof_threshold, m)
    duration = 0
    for k in range(max_duration + 1):
        product_term = 1
        for l in range(k):
            ho_failure_prob *= (l-1)
            product_term *= ho_failure_prob
        duration += ho_trigger_prob * product_term

    return duration
    
def calc_trigger_prob(gamma_T, m):
    # Calculate Beta function
    beta_function = gamma(m) * gamma(m)
    # Calculate hypergeometric function
    hypergeometric_func = hyp2f1(1, 2 * m, m + 1, 1 / (gamma_T + 1))
    probability = (gamma_T**m) / (m * (gamma_T + 1)**(2 * m)) * (beta_function**(-1)) * hypergeometric_func

    return probability

def calc_failure_prob(gamma_F, m):
    # Calculate Beta function
    beta_function = gamma(m) * gamma(m)
    probability = 1 - (gamma_F**m) / (m * (gamma_F + 1)**(2 * m)) * (beta_function**(-1)) * hyp2f1(1, 2 * m, m + 1, 1 / (gamma_F + 1))

    return probability

def set_connection(sat, gs):
    gs.set_current_satellite(sat)
    sat.set_current_gs(gs)
    # set_link_properties(net, )

def run(satellite_list, gs_list, total_cycle, output_file):
    gs_ho_info = {}
    # for gs in gs_list:
    #     gs.print_info()

    for curr_cycle in range(0, total_cycle):
        print(f"------------------------ Cycle {curr_cycle} / {total_cycle}------------------------")
        # Set candidate satellites for each ground station
        for gs in gs_list:
            if str(gs.id) not in gs_ho_info:
                gs_ho_info[str(gs.id)] = []
            ho_info = ['NULL', 'NULL', 'NULL']
            candidates = []
            for sat in satellite_list:
                sat_position = sat.get_position(curr_cycle)
                sat_position_cbf = lla2cbf(sat_position)    # Convert the satellite position (lat, long, alt) to coordinates (x,y,z)
                gs_position_cbf = lla2cbf(gs.get_position())     # Convert the ground station position (lat, long, alt) to coordinates (x,y,z)
                elevation = calculate_elevation_angle(sat_position_cbf, gs_position_cbf)
                L = getCoverageLimitL(elevation, sat_position_cbf[2])
                
                if checkSatCoverGroundStation(sat_position_cbf, gs_position_cbf, L):
                    candidates.append(sat)

            if not candidates:
                print(f"GS [{gs.id} {gs.name}] --> No candidate satellites at time {curr_cycle}")
                continue
            
            ## If current satellite is not equal to the candidate satellite, handover is processed
            if len(candidates) < 2:
                target_sat = candidates[0]
            else:
                target_sat, target_loss = select_next_satellite(candidates, gs.get_position(), curr_cycle)
            
            curr_sat = gs.get_current_satellite()
            if not curr_sat:
                ho_info = ['NULL', target_sat.id, 'NULL']
                set_connection(target_sat, gs)
            else: 
                if curr_sat != target_sat:
                    current_loss = calculate_link_loss(curr_sat.get_position(curr_cycle), gs.get_position())
                    # target_loss = calculate_link_loss(target_sat.get_position(curr_cycle), gs.get_position())
                    if current_loss / target_loss > ho_threshold:    
                        duration = calculate_freeze_duration()
                        ho_info = [curr_sat.id, target_sat.id, duration]
                        try:
                            # print(f"CWND is freezed during {duration} ms...")
                            # time.sleep(duration)
                            set_connection(target_sat, gs)
                        except Exception as exception:
                            print("Handover is failed...")

            gs_ho_info[str(gs.id)].append(ho_info)

    print(gs_ho_info)

    # # Convert string values to numpy array with dtype='object'
    # gs_ho_info_lists = {key: value if value == 'NULL' else list(map(str, value)) for key, value in gs_ho_info.items()}

    # # Save the dictionary with numpy arrays to the MAT file
    scio.savemat(output_file, gs_ho_info)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("need to specify the satellite position file!")
        exit()
    
    satellite_objs = []
    total_cycle = 0
    file_path = os.path.join(data_file_dir, sys.argv[1])
    try:
        satellite_positions = scio.loadmat(file_path)
        for id, sat_name in enumerate(satellite_positions):
            if 'STARLINK' in sat_name:
                sat = Satellite(id-2, sat_name, satellite_positions[sat_name])
                satellite_objs.append(sat)
                total_cycle = len(satellite_positions[sat_name])
    except Exception as exception:
        # Handle exceptions (you may want to log or print the exception)
        print(f"Error processing file {file_path}: {exception}")

    ## Generate ground station objects from 'ground_sations.xlsx' file
    gs_objs = []
    wb = openpyxl.load_workbook(os.path.join(data_file_dir, ground_station_file_name))
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):
        gs = Ground_station(i-2, sheet.cell(row = i, column = 3).value, float(sheet.cell(row = i, column = 1).value), float(sheet.cell(row = i, column = 2).value), 0)
        gs_objs.append(gs)
    
    ho_mat_file = './save/handover_info_file.mat'
    run(satellite_objs, gs_objs, total_cycle, ho_mat_file)

    print(f"Handover information is saved to '{ho_mat_file}'.")
